from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
import os
import tempfile
from datetime import datetime

app = Flask(__name__)
CORS(app)

FILE_NAME = "Pedidos.xlsx"

# Crear archivo Excel si no existe
if not os.path.exists(FILE_NAME):
    workbook = Workbook()
    workbook.save(FILE_NAME)


@app.route('/guardar_pedido', methods=['POST'])
def guardar_pedido():
    """Guarda un pedido en el archivo Excel"""
    data = request.json
    fecha = data.get("Fecha", "")

    if not fecha:
        return jsonify({"message": "Fecha inválida."}), 400

    workbook = load_workbook(FILE_NAME)
    if fecha not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=fecha)
        sheet.append(list(data.keys()))  # Agrega encabezados solo la primera vez
    else:
        sheet = workbook[fecha]

    sheet.append([data.get(key, "") for key in data.keys()])
    workbook.save(FILE_NAME)

    return jsonify({"message": "Pedido guardado correctamente"}), 200

@app.route('/consultar_pedidos', methods=['GET'])
def consultar_pedidos():
    """Consulta solo las hojas con registros y omite las vacías"""
    if not os.path.exists(FILE_NAME):
        return jsonify({"message": "No hay pedidos registrados."}), 404

    workbook = load_workbook(FILE_NAME)
    pedidos = {}

    for sheet in workbook.worksheets:
        headers = [cell.value for cell in sheet[1]]
        registros = [
            {headers[i]: row[i] for i in range(len(headers))} 
            for row in sheet.iter_rows(min_row=2, values_only=True) if any(row)
        ]

        if registros:  # Solo agrega la hoja si tiene registros
            pedidos[sheet.title] = registros

    return jsonify(pedidos) if pedidos else jsonify({"message": "No hay pedidos registrados."}), 200

@app.route('/actualizar_estatus', methods=['PUT'])
def actualizar_estatus():
    """Actualiza el estatus de un pedido y lo elimina si es 'Cancelado'"""
    data = request.json
    fecha, pedido_id, nuevo_estatus = data.get("Fecha"), data.get("ID"), data.get("Estatus")

    if not fecha or not pedido_id:
        return jsonify({"message": "Falta la fecha o el ID del pedido."}), 400

    workbook = load_workbook(FILE_NAME)
    sheet = workbook[fecha]

    for row in sheet.iter_rows(min_row=2):
        if str(row[0].value) == str(pedido_id):
            if nuevo_estatus == "Cancelado":
                sheet.delete_rows(row[0].row)
            else:
                row[7].value = nuevo_estatus  # Actualiza el estatus en la columna 7
            break

    workbook.save(FILE_NAME)
    return jsonify({"message": "Estatus actualizado correctamente"}), 200

@app.route('/eliminar_pedido', methods=['DELETE'])
def eliminar_pedido():
    """Elimina un pedido sin usar ID"""
    try:
        data = request.json
        fecha = data.get("Fecha")
        pedido_index = data.get("Index")

        if not fecha or pedido_index is None:
            return jsonify({"message": "Falta la fecha o el índice del pedido."}), 400

        workbook = load_workbook(FILE_NAME)
        sheet = workbook[fecha]

        sheet.delete_rows(int(pedido_index) + 2)  # Ajuste para la fila real en Excel
        workbook.save(FILE_NAME)

        return jsonify({"message": "Pedido eliminado correctamente."}), 200

    except Exception as e:
        return jsonify({"message": f"Error al eliminar el pedido: {str(e)}"}), 500
    
@app.route('/eliminar_pedidos_fecha', methods=['DELETE'])
def eliminar_pedidos_fecha():
    """Elimina todos los pedidos de una fecha específica"""
    try:
        data = request.json
        fecha = data.get("Fecha")

        if not fecha:
            return jsonify({"message": "Falta la fecha."}), 400

        workbook = load_workbook(FILE_NAME)
        
        if fecha in workbook.sheetnames:
            workbook.remove(workbook[fecha])
            workbook.save(FILE_NAME)
            return jsonify({"message": "Todos los pedidos de la fecha han sido eliminados."}), 200
        else:
            return jsonify({"message": "No se encontraron pedidos para la fecha dada."}), 404

    except Exception as e:
        return jsonify({"message": f"Error al eliminar los pedidos: {str(e)}"}), 500

@app.route('/descargar_excel', methods=['POST'])
def descargar_excel():
    """Descarga el archivo Excel con los pedidos"""
    workbook = load_workbook(FILE_NAME)
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    workbook.save(temp_file.name)
    temp_file.close()

    return send_file(temp_file.name, as_attachment=True, download_name="Pedidos.xlsx")

if __name__ == '__main__':
    from waitress import serve  # Alternativa a gunicorn en Windows
    serve(app, host='0.0.0.0', port=5000)